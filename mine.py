# file = open('transactions2.csv', 'r')


# data = file.read().splitlines()
# fline = data[0]

# print(fline)

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename1, filename2):
    wb = xl.load_workbook(filename1)
    sheet = wb['Sheet1']
    #cell = sheet['a1'] #sheet.cell(1,1)
    #print(cell.value)
    #print(sheet.max_row)

    for row in range(2, sheet.max_row + 1):
        print(row)
        cell1 = sheet.cell(row, 3)
        corrected_price = cell1.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4
              )
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    return wb.save(filename2)


# process_workbook('transactions2.xlsx', 'transactions3.xlsx')


wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

cell = sheet['a1'] 
dat = sheet.cell(1,1)

print(cell.value)
print(sheet.max_row)
print(sheet.max_column)
print('')

# sheetsname = wb.get_sheet_names()
# sht = wb.get_sheet_by_name('Sheet1')
# type(sht)
# sht['A1'].value
print('data sheets ' + dat)
# print(sheetsname)

#fir column 
# for i in range(1, sheet.max_column + 1):
#     cell_obj  = sheet.cell(1,i)
#     if cell_obj.value is not None:
#         print(cell_obj.value)

# print('')
# #first row
# for i in range(1, sheet.max_row + 1):
#     cell_obj1  = sheet.cell(i,1)
#     if cell_obj1.value is not None:
#         print(cell_obj1.value)


#
sheet['f1'].value  = 'category'
wb.save('new_transactions.xlsx')

sheet.title 
sheet.title  = 'My sheet name'
wb.save('new_transactions2.xlsx') 

sheet['f1'].value #sheet.cell(row-1, colunm-6).value
for i in range(1, 5): print(sheet.cell(row=i, column=3).value) 





xl.cell.get_column_letter(1) #get column letter, result would be 'A'
#inverse
xl.cell.column_index_from_string('A') #get column integer, result would be 1

wb.create_sheet(title="My sheet name", index=1)
wb.save('new_transactions3.xlsx')


sheet.column_dimensions('B').width = 20
sheet.row_dimensions(2).height = 70

wb.save('new_transactions4.xlsx')

From openpyxl.styles import Font

sheet['B1'].font = Font(sz=14, bold=True, italic=True)
wb.save('new_transactions5.xlsx')

wb = xl.Workbook()
sheet = wb.create_sheet('My Sheething')
import random 

for i in range(1, 11):
    sheet['A' + str(i)].value = random.randint(1, 100)

wb.save('new_transactions6.xlsx')


refObj = xl.chart.Reference(sheet, (1, 1), (10, 1))

#sheet['C8'].value = '=SUM(C1:C7)'

